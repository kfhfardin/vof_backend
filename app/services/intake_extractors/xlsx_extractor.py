"""Excel (.xlsx) extractor via openpyxl.

Iterates every sheet, returns each sheet's rows as header-keyed dicts.
The `rows` field aggregates all sheets; per-sheet structure is in metadata.
"""

import io
from typing import Any, BinaryIO, ClassVar

from openpyxl import load_workbook

from app.services.intake_extractors.base import ExtractedContent, IntakeExtractor


class XlsxExtractor(IntakeExtractor):
    name: ClassVar[str] = "xlsx"
    accepts_mime: ClassVar[list[str]] = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ]
    accepts_ext: ClassVar[list[str]] = [".xlsx", ".xlsm"]

    def extract(self, blob: BinaryIO, filename: str) -> ExtractedContent:
        bio = io.BytesIO(blob.read())
        wb = load_workbook(bio, read_only=True, data_only=True)
        all_rows: list[dict[str, Any]] = []
        sheets_meta: list[dict[str, Any]] = []
        try:
            for sheet in wb.worksheets:
                rows_iter = sheet.iter_rows(values_only=True)
                try:
                    header_row = next(rows_iter)
                except StopIteration:
                    sheets_meta.append({"name": sheet.title, "row_count": 0})
                    continue
                headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header_row)]
                sheet_rows: list[dict[str, Any]] = []
                for row in rows_iter:
                    sheet_rows.append(
                        {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
                    )
                all_rows.extend(sheet_rows)
                sheets_meta.append({"name": sheet.title, "row_count": len(sheet_rows)})
        finally:
            wb.close()
        return ExtractedContent(
            rows=all_rows,
            metadata={"sheets": sheets_meta, "row_count": len(all_rows)},
        )
